##
# excelToJson
# AUTHOR: MARCO SELVA
# DATE CREATION 02/12/2022
# VERSION: 2.0
# v2.0: 01/12/2023
# v1.0: 03/06/2022
# 
# TODO
# 
# ##

from openpyxl import load_workbook
from datetime import datetime, timedelta
from json import dump
import glob, os, requests, sys

def log(*args, end='\n'):
    global logger
    print(*args, end=end)
    logger.write("".join(args)+end)

def getStrategy(name):
    default = '6220e21a9344202f70a26818'
    if not name:
        return default
    r = requests.post('http://95.111.245.75/api/report/strategyInfoByName', json={'name': name})
    if not r or 'error' in r.text:
        return default
    return (r.json() or [{}])[0].get('id', default)


def getRunnerId(runner):
    name = runner['name']
    if default := runner['id']:
        return default
    r = requests.post('http://95.111.245.75/api/runners/infoByName', json={'name': name})
    if not r or 'error' in r.text:
        return default
    return (r.json() or [{}])[0].get('id', default)


def getStamp(date, time):
    try:
        if not time:
            return None
        if isinstance(time, str):
            time = datetime.strptime(time.replace('_', ''), '%I:%M:%S %p' if 'm' in time else '%H:%M:%S' )
        return int(datetime.strptime(date.strftime('%Y-%m-%d ') + time.strftime('%H:%M:%S'),
                                     '%Y-%m-%d %H:%M:%S').timestamp() * 1000) + 3600000
    except AttributeError:
        log("\t [Possible Error: Date not present]", end='')
        return None


def getAvg(block, runner, char, type_): #fixed
    stake = 0
    oddsWeight = 0
    for row in block:
        if row[30] == runner['char'] and row[31] == char:
            stake += row[33] or 0
            oddsWeight += (row[32] or 0) * (row[33] or 0)
    odds = 0 if stake == 0 else (oddsWeight / stake)
    if type_ == 'back':
        toWin = stake * (odds - 1)
        liability = stake
    else:
        liability = stake * (odds - 1)
        toWin = stake
    return { # fixed
        'odds': round(odds, 2),
        'stake': round(stake, 2),
        'toWin': round(toWin, 2),
        'liability': round(liability, 2)
    }


def checkTime(time):
    if isinstance(time, datetime):
        time = time.time()
    return time


def getTime(rn, rows):
    global day
    time = checkTime(rows[rn][28])
    date = rows[-1][1]
    prevTime = time
    if rn < len(rows) - 1:
        prevTime = checkTime(rows[rn + 1][28])
    if not prevTime:
        return getStamp(rows[-1][1] - timedelta(day), time)
    if not time:
        log("\t [Possible Error: Empty Time in sets/trades]", end='')
        return None
    elif day or prevTime > time:
        date = rows[-1][1] - timedelta(day)
        day = (day + 1) if prevTime > time else 1
    return getStamp(date, time)


def getStat(rn, block):
    row = block[0][rn]
    return row if row not in ['#N/A', None, '0'] else 0


def compileData(b, block, inc):
    global day, outputPath
    log(f"\t[{b}] {block[0][3]}", end="")
    if block[-1][14] == None:
        log("\t => [Error: Incomplete sets]")
        return
    json = dict(
        created=int(datetime.now().timestamp() * 1000) + 3600000,
        updated=int(datetime.now().timestamp() * 1000) + 3600000,
        trade=dict(
            info=dict(
                setTime=dict(
                    # check if remove 1 hour in all from next
                    second=getStamp(block[0][1], block[1][2]),
                    third=getStamp(block[0][1], None if len(block) < 3 else block[2][2])
                ),
                strategyId=getStrategy(block[0][6]),
                tennisTournamentId=block[1][7],
                date = getStamp(block[0][1],block[0][2] or '00:00:00'), 
                marketInfo=dict(
                    marketName=block[0][3],
                    marketId='',
                    # if "-" not present default is MATCH_ODDS, else use the suffix after " - "
                    marketType=(
                        "MATCH_ODDS" if ' - ' not in block[0][3] else block[0][3].split(' - ')[-1].
                            replace(' 1 ', '_').replace(' 2 ', '_')).upper(),
                    eventName="Match Odds" if ' - ' not in block[0][3] else block[0][3].split(' - ')[-1],
                    sport="TENNIS"
                ),
                executor=[block[0][5]],
                exchange=dict(
                    name=block[0][4],  # COLUMN E
                    commission=0.02 if block[0][4] in ['MT KEVIN',' UK', 'DEMO'] else 0.05,
                ),
                note=dict(
                    description='',
                    entry='\n'.join([row[13] or '' for row in block]).strip(),
                    exit='',
                    position='',
                    mm='',
                    odds='',
                    post=''
                )
            ),
            selections=[],
            trades=[],
            results=dict(),
            stats=[],
            params=[]
        ),
    )
    runners = [
        dict(
            name=block[0][8 + runner],
            bsp=block[1][8 + runner],
            id=(None if len(block) < 3 else block[2][8 + runner]),
            sets=dict(
                secondSet=block[0][11 + runner],
                thirdSet=block[1][11 + runner],
            ),
            char=['A', 'B'][runner]
        )
        for runner in range(2)]

    json['trade']['selections'] = [dict(
        selectionN=rn,
        runnerId=getRunnerId(runner),
        runnerName=runner['name'],
        winner=block[0][10] == runner['name'],
        bsp=runner['bsp'],
        sets=runner['sets'],
        avg=dict(
            back=getAvg(block, runner, 'B',"back"),
            lay=getAvg(block, runner, 'L',"lay"), # fixed
        )
    ) for rn, runner in enumerate(runners)]
    day = 0
    if (time:=checkTime(block[-2][28])) and (prevTime:=checkTime(block[0][2])) and time > prevTime:
        day = 1 
    json['trade']['trades'] = [dict(
              id=len(block[:-1]) - rn,
              selectionN=['A', 'B', None].index(row[30]),
              type=['back', 'lay'][(type_ := row[31]) == 'L'],
              odds=round((odds := float(row[32] or 0)), 2),
              stake=round((stake := float(row[33] or 0)), 2),
              liability=round((stake if type_ == 'B' else (stake * (odds - 1))), 2),
              ifWin=round(((stake * (odds - 1)) if type_ == 'B' else stake), 2),
              options=row[29],
              condition=dict(
                  tennis=dict(
                      isTennis=True,
                      point=dict(
                          set1=dict(
                              runnerA=row[14] or 0,
                              runnerB=row[15] or 0,
                          ),
                          set2=dict(
                              runnerA=row[16] or 0,
                              runnerB=row[17] or 0,
                          ),
                          set3=dict(
                              runnerA=row[18] or 0,
                              runnerB=row[19] or 0,
                          ),
                          set4=dict(
                              runnerA=row[20] or 0,
                              runnerB=row[21] or 0,
                          ),
                          set5=dict(
                              runnerA=row[22] or 0,
                              runnerB=row[23] or 0,
                          ),
                          currentGame=dict(
                              runnerA=str(row[24] or 0),
                              runnerB=str(row[25] or 0),
                              server=str(row[26] or "")
                          )
                      )
                  ),
                  football=dict(
                      isFootball=False,
                      point=dict(
                          home=0,
                          away=0,
                      )
                  ),
                  horse=dict(
                      isHorse=False,
                  ),
                  note=row[13],
                  time=getTime(rn, block[:-1][::-1]),
              )
          )
          for rn, row in enumerate(block[:-1][::-1])][::-1]
    
    tradeIndex = 0
    trades = []
    for trade in json['trade']['trades']:
        if not trade['odds']:
            log("\t [Error -> Empty row skipped]")
            continue
        tradeIndex += 1            
        trade['id'] = tradeIndex
        trades.append(trade)
    json['trade']['trades'] = trades

    json['trade']['results'] = dict(
        grossProfit=round((aj := float(block[0][35] or 0)) + (ak := float(block[0][36] or 0)), 2),  # rounded
        netProfit=round(ak, 2),  # rounded
        maxRisk=round(-(ai := float(block[0][34] or 0)), 2),  # rounded
        rr=0 if ai == 0 else round(ak / ai, 4),  # 4 digits
        commissionPaid=aj,
        correctionPl=0,
        finalScore=dict(
            tennis=dict(
                set1=dict(
                    runnerA=block[-1][14],
                    runnerB=block[-1][15],
                ),
                set2=dict(
                    runnerA=block[-1][16],
                    runnerB=block[-1][17],
                ),
                set3=dict(
                    runnerA=block[-1][18],
                    runnerB=block[-1][19],
                ),
                set4=dict(
                    runnerA=block[-1][20],
                    runnerB=block[-1][21],
                ),
                set5=dict(
                    runnerA=block[-1][22],
                    runnerB=block[-1][23],
                ),
                currentGame=dict(
                    runnerA=None,
                    runnerB=None,
                    server=None
                )
            ),
            football=dict(
                home=0,
                away=0,
            )
        )
    )
    json['trade']['stats'] = [
        dict(
            runnerId=json['trade']['selections'][[0, 1][54 == st]]['runnerId'],
            stats1=getStat(st + 0, block),
            stats2=getStat(st + 1, block),
            stats3=getStat(st + 2, block),
            stats4=getStat(st + 3, block),
            stats5=getStat(st + 4, block),
            stats6=getStat(st + 5, block),
            stats7=getStat(st + 6, block),
            stats8=getStat(st + 7, block),
            stats9=getStat(st + 8, block),
            stats10=getStat(st + 9, block),
        )
        for st in range(44, 55, 10)]

    json['trade']['params'] = [
        dict(
            runnerId=json['trade']['selections'][[0, 1][75 == st]]['runnerId'],
            params1=block[0][st + 0] or 0,
            params2=block[0][st + 1] or 0,
            params3=block[0][st + 2] or 0,
            params4=block[0][st + 3] or 0,
            params5=block[0][st + 4] or 0,
            params6=block[0][st + 5] or 0,
            params7=block[0][st + 6] or 0,
            params8=block[0][st + 7] or 0,
            params9=block[0][st + 8] or 0,
            params10=block[0][st + 9] or 0,
        )
        for st in range(65, 76, 10)]
    filename = block[0][1].strftime(f'{inc}-%m_%d_%Y_{block[0][3]}.json').replace('/','_')
    log(f"\t => [Compiled to {filename}]")
    dump(json, open(os.path.join(outputPath, filename), 'w', encoding='utf-8'), indent=4)


inputPath = "Input"
outputPath = "Output"

outputPath = os.path.join(outputPath, datetime.now().strftime('%m_%d_%Y_%H_%M_%S'))
if not os.path.exists(outputPath):
    try:
        os.mkdir(outputPath)
    except:
        log("[-] Output Path not exists and unable to Create")
        sys.exit(0)

logger = open(os.path.join(outputPath,'00_logs.txt'), 'w', encoding='utf-8')

for file in glob.glob(os.path.join(inputPath, "*.xlsx")):
    log(f'[=] Reading Book: {os.path.basename(file)}')
    try:
        book = load_workbook(file)
    except Exception as e:
        log(f"[-] Unable to read Book: {os.path.basename(file)}, Error: {e}")
        continue
    sheet = book.active
    block = []
    data = []
    for row in sheet.iter_rows():
        cells = [cell.value for cell in row]
        if 'OPEN' in cells:
            block.append(cells)
        elif 'FINAL' in cells:
            block.append(cells)
            data.append(block)
            block = []
        elif block:
            block.append(cells)
    inc = 1
    for b, block in enumerate(data, start=1):
        try:
            compileData(b, block, inc)
            inc += 1
        except Exception as e:
            log(f" => [Error: {e}]")
    log()
logger.close()