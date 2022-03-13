from openpyxl import load_workbook
from datetime import datetime, timedelta
from json import dump
import glob, os, requests, sys


def getStrategy(name):
    default = '6220e21a9344202f70a26818'
    if not name:
        return default
    r = requests.post('http://217.61.104.122/api/report/strategyInfoByName',json={'name':name})
    
    if not r or 'error' in r.text:
        return default
    return (r.json() or [{}])[0].get('id',default)

def getRunnerId(runner):
    name = runner['name']
    if default := runner['id']:
        return default
    r = requests.post('http://217.61.104.122/api/runners/infoByName',json={'name':name})
    if not r or 'error' in r.text:
        return default
    return (r.json() or [{}])[0].get('id',default)

def getStamp(date,time):
    if isinstance(time,str) or not time:
        time = datetime.strptime((time or '00:00:00').replace('_',''),'%H:%M:%S')
    return int(datetime.strptime(date.strftime('%Y-%m-%d ') + (time).strftime('%H:%M:%S'),'%Y-%m-%d %H:%M:%S').timestamp())

def getAvg(block,runner,char,params=[]):
    stake = sum([row[33] or 0 for row in block if row[30]==runner['char'] and row[31]==char])
    odds = sum([row[32] or 0 for row in block if row[30]==runner['char'] and row[31]==char])
    odds = 0 if stake == 0 else ((stake * odds)/stake) # the Weighted arithmetic mean for each bets B "back" (column AF), its like: odds = (summation of odds (column AG) * stake (columns AH) ) / (sum of the stake)

    stake = stake # the sum of stake bets B "back" (column AF) over the runner (colum AE to select which ones)
    toWin = stake * (odds -1 ) # toWin = stake * (odds-1)
    return {
        params[0]:round(odds,2),
        params[1]:round(stake,2),
        params[2]:round(toWin,2)
    }
def checkTime(time):
    if isinstance(time,datetime):
        time = time.time()
    return time
def getTime(rn,rows):
    global day
    time = checkTime(rows[rn][28])
    date = rows[-1][1]
    prevTime = time
    if rn < len(rows)-1:
        prevTime = checkTime(rows[rn+1][28])
    if not prevTime:
        return getStamp(rows[-1][1] - timedelta(day), time)
    if not time:
        return None
    elif day or prevTime>time:
        date = rows[-1][1] - timedelta(day)
        day = (day+1) if prevTime>time else 1
    
    return getStamp(date,time)
def getStat(rn,block):
    return sum([(row[rn] if row[rn] not in ['#N/A',None,'0'] else 0) for row in block])

def compileData(data):
    global day, outputPath
    for b, block in enumerate(data,start=1):
        print(f"\t[{b}] {block[0][3]}",end="")
        if not block[-1][14]:
            print(" => [Error: Incomplete sets]")
            continue
        json = dict(
            created = int(datetime.now().timestamp()),
            updated = int(datetime.now().timestamp()),
            trade = dict(
            # for that part take attention of columns from B to N
            info = dict(
                strategyId = getStrategy(block[0][6]), # explained in API STRATEGY paragraph
                tennisTournamentId = block[1][7], # the ID in second row of each columns H "event"
                date = getStamp(block[0][1],block[0][2]),   # UTC millisecond timestamp of the market (get date in column A and time in column C)
                marketInfo = dict(
                    marketName = block[0][3], # market columns D
                    marketId = '',   # leave empty for the moment
                    marketType = ("MATCH_ODDS" if ' - ' not in block[0][3] else block[0][3].split(' - ')[-1].replace(' 1 ','_').replace(' 2 ','_')).upper(),  
                    eventName = "Match Odds" if ' - ' not in block[0][3] else block[0][3].split(' - ')[-1],
                    sport = "TENNIS"
                ),
                executor = [block[0][5]],     
                exchange = dict(
                    name = block[0][4],       # COLUMN E
                    commission = 0.02 if block[0][4]=='UK' else 0.05, 
                ),
                note = dict(
                    description = '', 
                    entry = '\n'.join([row[13] or '' for row in block]).strip(), # merge all the comments here from column N
                    exit = '', 
                    position = '',
                    mm = '', 
                    odds = '', 
                    post = ''
                )
            ),
            ))
        runners = [
            dict(
                name=block[0][8+runner],
                bsp=block[1][8+runner],
                id=(None if len(block)<3 else block[2][8+runner])
                ,
                sets = dict(
                    secondSet = block[0][11+runner], # column L for runnerA, columns for runner B, the first row
                    thirdSet = block[1][11+runner], # column L for runnerA, columns for runner B, the second row
                ),
                char = ['A','B'][runner]
            ) 
            for runner in range(2)]

        json['trade']['selections'] = [dict( # array of objet, one each for runner (runnerA in columns I and runnerB in column J)
                selectionN = rn, # sequential number of runner, 0 for runnerA, 1 for runnerB
                runnerId = getRunnerId(runner),   #// explained in API RUNNERS paragraph
                runnerName = runner['name'], # the name of the runner, first row in column I or J
                winner = block[0][10]==runner['name'], #// if the runnerName is the same name in column K true, false otherwise
                bsp = runner['bsp'], # the second row of the columns of runner I or J
                sets = runner['sets'],
                avg = dict(
                    back = getAvg(block,runner,'B',['odds','stack','toWin']),
                    lay = getAvg(block,runner,'L',['adds','bank','liability'])
                )
        ) for rn, runner in enumerate(runners)]
        day = 0
        json['trade']['trades'] = [ # for that part take attention of columns from N to AK, now we merge matched bets and point to have single trade for each row
                dict(
                    id = row[27], #  the seq id, column AB
                    selectionN = ['A','B',None].index(row[30]), # sequential number of runner (0 for runnerA, 1 for runnerB)
                    type = ['back','lay'][(type:=row[33])=='L'], #  columns AF ->if B "back" , if L "lay"
                    odds = round((odds:=float(row[32] or 0)),2), #  columns AG
                    stake = round((stake:=float(row[33] or 0)),2), #  colum AK
                    liability = round((stake if type=='B' else (stake*(odds-1))),2), #  if type=="B" so liability = stake (column AK), if "L" : liability = stake*(odds-1)
                    ifWin = round(((stake*(odds-1)) if type=='B' else stake),2), #  if type=="B" ifWin = stake*(odds-1), if "L" : ifWin = stake (columns AK)
                    options = row[29], #// columns AD
                    condition = dict(
                        tennis = dict(
                            isTennis = True, # true always
                            point = dict( # 0 if the cell are empty
                                set1 = dict(
                                    runnerA = row[14] or 0, #// column O
                                    runnerB = row[15] or 0, #// column P
                                ),
                                set2 = dict(
                                    runnerA = row[16] or 0, #// column Q
                                    runnerB = row[17] or 0, #// column R
                                ),
                                set3 = dict(
                                    runnerA = row[18] or 0, #// column S
                                    runnerB = row[19] or 0, #// column T
                                ),
                                set4 = dict(
                                    runnerA = row[20] or 0, # column U
                                    runnerB = row[21] or 0, # column V
                                ),
                                set5 = dict(
                                    runnerA = row[22] or 0, # column W
                                    runnerB = row[23] or 0, # column X
                                ),
                                currentGame = dict(
                                    runnerA = row[24] or 0, # column Y
                                    runnerB = row[25] or 0, # column Z
                                    server = row[26] # column AA
                                )
                            )
                        ),
                        football = dict( 
                            isFootball = False, # false
                            point = dict(
                                home = 0, #// 0 
                                away = 0, #// 0 
                            )
                        ),
                        horse = dict( 
                            isHorse = False, # false
                        ),
                        note = row[13], # column N
                        time = getTime(rn,block[:-1][::-1]), # UTC millisecond timestamp of the bets, column AC, (first date is the same of market, if have bets in different day, such as 23:34:46 and after 00:12:46 so the second have day = firstdate +1)
                    )
                )
            for rn, row in enumerate(block[:-1][::-1])][::-1]
        json['trade']['results'] =  dict(
                grossProfit = (aj:=float(block[0][35] or 0)) + (ak:=float(block[0][36] or 0)), # column AJ + column AK
                netProfit = ak, # column AK
                maxRisk = -(ai:=float(block[0][34] or 0)), # -column AI
                rr = 0 if ai==0 else round(ak/ai,2), # column AK/column AI
                commissionPaid = aj, # column AJ
                correctionPl = 0, # 0
                finalScore = dict( # from the row where "FINAL" are in column AD, get the point at here left, column 0
                    tennis = dict( # 0 if the cell are empty
                         set1 = dict(
                            runnerA = block[-1][14], # column O
                            runnerB = block[-1][15], # column P
                        ),
                        set2 = dict(
                            runnerA = block[-1][16], # column Q
                            runnerB = block[-1][17], # column R
                        ),
                        set3 = dict(
                            runnerA = block[-1][18], # column S
                            runnerB = block[-1][19], # column T
                        ),
                        set4 = dict(
                            runnerA = block[-1][20], # column U
                            runnerB = block[-1][21], # column V
                        ),
                        set5 = dict(
                            runnerA = block[-1][22], # column W
                            runnerB = block[-1][23], # column X
                        ),
                        currentGame = dict(
                            runnerA = 0,
                            runnerB = 0,
                            server = None
                        )
                     ),
                    football = dict(
                        home = 0,
                        away = 0,
                    )
                )
            )
        json['trade']['stats'] = [
                dict(
                    runnerId = json['trade']['selections'][[0,1][54==st]]['runnerId'], # the runnerId, taken form column I or J third row if present, via API if not present
                    stats1 = getStat(st+0,block), # runnerA column AS, runnerB column BC
                    stats2 = getStat(st+1,block), # runnerA column AT, runnerB column BD
                    stats3 = getStat(st+2,block), # runnerA column AU, runnerB column BE
                    stats4 = getStat(st+3,block), # runnerA column AV, runnerB column BF
                    stats5 = getStat(st+4,block), # runnerA column AW, runnerB column BG
                    stats6 = getStat(st+5,block), # runnerA column AX, runnerB column BH
                    stats7 = getStat(st+6,block), # runnerA column AY, runnerB column BI
                    stats8 = getStat(st+7,block), # runnerA column AZ, runnerB column BJ
                    stats9 = getStat(st+8,block), # runnerA column BA, runnerB column BK
                    stats10 = getStat(st+9,block), # runnerA column BB, runnerB column BL
              )
                for st in range(44,55,10)]

        filename = block[0][1].strftime(f'%m_%d_%Y_{block[0][3]}.json')
        
        print(f" => [Compiled to {filename}]")
        dump(json,open(os.path.join(outputPath,filename),'w',encoding='utf-8'),indent=4)








inputPath = "InputFolder"
outputPath = "OutputFiles"

outputPath = os.path.join(outputPath,datetime.now().strftime('%m_%d_%Y_%H_%M_%S'))
if not os.path.exists(outputPath):
    try:
        os.mkdir(outputPath)
    except:
        print("[-] Output Path not exists and unable to Create")
        sys.exit(0)


for file in glob.glob(os.path.join(inputPath,"*.xlsx")):
    print(f'[=] Reading Book: {os.path.basename(file)}')
    try:
        book = load_workbook(file)
    except Exception as e:
        print(f"[-] Unable to read Book: {os.path.basename(file)}, Error: {e}")
        continue
    sheet = book.active
    block = []
    data = []
    for row in sheet.iter_rows():
        cells = [cell.value for cell in row]
        if 'OPEN' in cells:
            block.append(cells)
        elif 'FINAL' in cells:
            block.append(cells)
            data.append(block)
            block = []
        elif block:
            block.append(cells)
    try:
        compileData(data)
    except ValueError as e:
        print(f" => [Error: {e}]")
    print()

            